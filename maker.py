import openpyxl
from Vlookup_grz.Reqests import Reqests


def maker():
    ses = Reqests()
    lms = ses.get_lm_grz()
    grz_in_lm = []

    for index, lm in enumerate(lms.get("rules")):
        print(index)
        grzs = ses.get_all_grz(lm.get("id"))
        grz_list = []
        try:
            for grz in grzs.get("conditions")[0].get("numbers"):
                grz_list.append(grz)
        except:
            pass
        grz_in_lm.append([lm.get("title"), grz_list])
    adress_list = ["Кубинка", "Волоколамское", "Приречная", "Шипиловский", "Ленинцев", "Шишкин", "Солнечная",
                   "Сельскохозяйственная", "Василевского", "Мосфильмовская", "Наличная", "Солдатская", "Тухачевского",
                   "Федосьино", "Шепелюгинская", "Багрицкого", "Миля", "Смольная", "Саратовская", "Цеткин",
                   "Климентовский"]
    Wordbook = openpyxl.load_workbook("ГРЗ_дворы.xlsx")
    sheet = Wordbook['result']
    sheet.cell(1, 5).value = ""
    for i in range(2, sheet.max_row + 1):
        print(i)
        target_adress = ""
        for adress in adress_list:
            if adress in str(sheet.cell(i, 3).value):
                target_adress = adress
                break
        for j in grz_in_lm:
            if target_adress in j[0] and str(sheet.cell(i, 5).value) in j[1]:
                sheet.cell(i, 8).value = "Есть"
                break
            sheet.cell(i, 8).value = "Нет"
    Wordbook.save(f'ГРЗ_итоги.xlsx')



